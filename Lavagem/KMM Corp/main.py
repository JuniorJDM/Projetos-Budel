import keyboard
import pyautogui
import openpyxl
from openpyxl.styles import PatternFill
from time import sleep, time
import subprocess
import win32gui
import win32con
# Fim bibliotecas

def bring_window_to_front_and_maximize(window_title):
    hwnd = win32gui.FindWindow(None, window_title)
    if hwnd:
        win32gui.SetForegroundWindow(hwnd)
        win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)


def carregar_planilha(caminho_arquivo_excel, planilha_nome):
    workbook = openpyxl.load_workbook(caminho_arquivo_excel)
    planilha = workbook[planilha_nome]
    return workbook, planilha


def carregar_linhas_processadas(caminho_arquivo):
    try:
        with open(caminho_arquivo, 'r') as f:
            linhas_processadas = [int(linha.strip()) for linha in f]
    except FileNotFoundError:
        linhas_processadas = []
    return linhas_processadas


def salvar_linhas_processadas(caminho_arquivo, linhas_processadas):
    with open(caminho_arquivo, 'w') as f:
        for linha in linhas_processadas:
            f.write(str(linha) + '\n')


def salvar_linhas_nao_processadas(caminho_arquivo, row_number, dados_linha):
    with open(caminho_arquivo, 'a') as f:
        f.write(f"Dados da linha {row_number}: {', '.join(map(str, dados_linha))}\n")


def reiniciar_kmm():
    subprocess.run(['taskkill', '/f', '/im', 'javaw.exe'])
    sleep(10)
    keyboard.press_and_release('win+r')
    sleep(5)
    keyboard.press('delete')
    sleep(5)
    keyboard.write('cmd')
    sleep(5)
    keyboard.press('enter')
    sleep(5)
    keyboard.write('start C:\\Users\\aprendiz.pcm\\Desktop\\KMM.lnk')
    sleep(5)
    keyboard.press('enter')
    sleep(5)
    subprocess.run(['taskkill', '/f', '/im', 'cmd.exe'])
    pyautogui.click(708, 399, duration=5)
    pyautogui.doubleClick(708, 399, duration=5)
    keyboard.press('delete')
    keyboard.write('ANTONIO.JUNIOR')
    pyautogui.click(602, 427, duration=5)
    keyboard.write('18117702')
    sleep(5)
    keyboard.press('enter')
    sleep(5)
    window_title = "KMM Corporativo :: 76.667.682 - BUDEL TRANSPORTES LTDA v3.23.1 Corporate"
    bring_window_to_front_and_maximize(window_title)
    sleep(5)
    keyboard.write('csc')
    sleep(5)
    keyboard.press('enter')
    sleep(5)
    pyautogui.click(1113, 15, duration=5)


def processar_linha(linha):
    TITULO = linha[0].value
    CUSTO = linha[1].value
    AREA = linha[2].value
    TABELA = linha[3].value
    PLACA = linha[4].value
    DATA = linha[5].value
    PRODUTO = linha[6].value
    TOTAL = linha[7].value

    pyautogui.click(35, 40, duration=0.2)
    pyautogui.click(60, 100, duration=0.2)
    pyautogui.click(134, 159, duration=0.2)
    keyboard.write(TITULO)
    pyautogui.click(179, 295, duration=0.2)
    keyboard.write(str(CUSTO))
    keyboard.press('tab')
    pyautogui.click(212, 344, duration=0.2)
    pyautogui.doubleClick(212, 344, duration=0.2)
    keyboard.press('delete')
    keyboard.write("066007")
    keyboard.press('tab')
    pyautogui.click(191, 371, duration=0.2)
    keyboard.write(str(AREA))
    keyboard.press('tab')
    pyautogui.click(368, 400, duration=0.2)
    keyboard.write(str(TABELA))
    keyboard.press('tab')
    pyautogui.click(72, 535, duration=0.2)
    pyautogui.click(472, 277, duration=0.2)
    keyboard.write(PLACA)
    pyautogui.click(476, 444, duration=0.2)
    keyboard.write(DATA)
    keyboard.press('tab')
    pyautogui.click(761, 543, duration=0.2)
    pyautogui.click(587, 536, duration=0.2)
    pyautogui.click(129, 98, duration=0.2)
    pyautogui.click(65, 287, duration=0.2)
    pyautogui.click(205, 131, duration=0.2)
    pyautogui.click(538, 177, duration=0.2)
    pyautogui.click(474, 231, duration=0.2)
    pyautogui.click(615, 177, duration=0.2)
    keyboard.write(PRODUTO)
    pyautogui.click(607, 180, duration=0.2)
    keyboard.press('backspace')
    pyautogui.click(959, 177, duration=0.2)
    pyautogui.click(719, 240, duration=0.2)
    pyautogui.click(832, 614, duration=0.2)
    pyautogui.click(170, 193, duration=0.2)
    pyautogui.typewrite('1')
    pyautogui.click(567, 194, duration=0.2)
    keyboard.write(str(TOTAL))
    keyboard.press('tab')
    pyautogui.click(520, 250, duration=0.2)
    keyboard.write('l')
    keyboard.press('tab')
    pyautogui.click(318, 281, duration=0.2)
    pyautogui.click(1096, 702, duration=0.2)
    sleep(2)

    return TITULO, CUSTO, AREA, TABELA, PLACA, DATA, PRODUTO, TOTAL


def verificar_imagem_e_confirmar(tempo_max_espera):
    tempo_inicial_verificacao = time()
    while time() - tempo_inicial_verificacao < tempo_max_espera:
        if pyautogui.locateOnScreen('C:\\Users\\aprendiz.pcm\\Desktop\\site\\confirmar.png', confidence=0.8) is not None:
            pyautogui.click(883, 421, duration=0.2)
            sleep(1)
            return True
        sleep(1)
    return False


def iniciar_codigo():
    print("Código iniciado!")
    keyboard.wait("F")
    sleep(5)
    caminho_arquivo_excel = 'C:\\Users\\aprendiz.pcm\\OneDrive\\Projetos_V2\\S.c_Corp\\formato.xlsx'
    planilha_nome = 'Plan1'
    workbook, planilha = carregar_planilha(caminho_arquivo_excel, planilha_nome)
    tempo_max_espera = 10
    linhas_processadas = carregar_linhas_processadas('linhas_processadas.txt')
    tentativas_por_linha = {}

    row_number = 2
    while row_number <= planilha.max_row:
        if row_number in linhas_processadas:
            row_number += 1
            continue

        linha = planilha[row_number]

        if len(linha) >= 8:
            dados_linha = processar_linha(linha)
            if verificar_imagem_e_confirmar(tempo_max_espera):
                linhas_processadas.append(row_number)
                salvar_linhas_processadas('linhas_processadas.txt', linhas_processadas)
                tentativas_por_linha.pop(row_number, None)
                row_number += 1
            else:
                print(f"Imagem não encontrada na linha {row_number}.")
                if row_number in tentativas_por_linha:
                    tentativas_por_linha[row_number] += 1
                else:
                    tentativas_por_linha[row_number] = 1

                if tentativas_por_linha[row_number] == 1:
                    salvar_linhas_nao_processadas('linhas_nao_processadas.txt', row_number, dados_linha)

                reiniciar_kmm()
                row_number += 1
        else:
            row_number += 1

    workbook.save(caminho_arquivo_excel)
    workbook.close()


# Chamar a função para iniciar o código
iniciar_codigo()
