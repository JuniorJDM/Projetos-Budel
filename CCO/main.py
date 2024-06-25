import keyboard
import pyautogui
import openpyxl
from openpyxl.styles import PatternFill
from time import sleep, time
import subprocess
import win32gui
import win32con

def bring_window_to_front_and_maximize(window_title):
    hwnd = win32gui.FindWindow(None, window_title)
    if hwnd:
        win32gui.SetForegroundWindow(hwnd)
        win32gui.ShowWindow(hwnd, win32con.SW_SHOWMAXIMIZED)

def iniciar_codigo():
    print("Código iniciado!")

    # Aguarda pela tecla F ser pressionada
    keyboard.wait("F")

    sleep(5)
    # Abre o arquivo Excelcsc
    caminho_arquivo_excel = 'O:\\AUTOMAÇÕES\\Dados\\CCO\\Pedido_Compra.xlsx'
    planilha_nome = 'Plan1'  # Nome da planilha que você deseja ler
    workbook = openpyxl.load_workbook(caminho_arquivo_excel)
    planilha = workbook[planilha_nome]

    # Define o tempo máximo de espera para a imagem aparecer (em segundos)
    tempo_max_espera = 10

    # Carrega ou cria o arquivo de registro de linhas processadas
    try:
        with open('linhas_processadas.txt', 'r') as f:
            linhas_processadas = [int(linha.strip()) for linha in f]
    except FileNotFoundError:
        linhas_processadas = []

    # Dicionário para armazenar o número de tentativas de cada linha
    tentativas_por_linha = {}

    # Itera sobre as linhas da planilha
    row_number = 2  # Começa da segunda linha
    while row_number <= planilha.max_row:
        # Verifica se a linha já foi processada
        if row_number in linhas_processadas:
            row_number += 1
            continue

        linha = planilha[row_number]

        # Certifique-se de que há elementos suficientes antes de tentar acessar os índices
        if len(linha) >= 8:
            #data, valor, lavagem, placa, cnpj = linha
            TITULO = linha[0].value
            CUSTO = linha[1].value
            RESPONSAVEL = linha[2].value
            AREA = linha[3].value
            TABELA = linha[4].value
            PLACA = linha[5].value
            PRODUTO = linha[6].value
            TOTAL = linha[7].value
            DATA = linha[8].value

            ## código de automação ##

            pyautogui.click(35, 40, duration=0.2)
            pyautogui.click(60, 100, duration=0.2)

            # Escrever
            pyautogui.click(134, 159, duration=0.2)
            pyautogui.write(TITULO)

            # Incluir centro de custo
            pyautogui.click(179, 295, duration=0.2)
            pyautogui.write(str(CUSTO))
            keyboard.press('tab')

            ##RESPONSAVEL
            pyautogui.click(212,344,duration=0.2)
            pyautogui.doubleClick(212,344,duration=0.2)
            keyboard.press('delete')
            pyautogui.write(RESPONSAVEL)
            keyboard.press('tab')

            # Área
            pyautogui.click(191, 371, duration=0.2)
            pyautogui.write(str(AREA))
            keyboard.press('tab')

            # Tabela de preço
            pyautogui.click(238,427, duration=0.2)
            pyautogui.write(TABELA)
            pyautogui.click(660,238,duration=0.2)
            pyautogui.click(823,613,duration=0.2)
            # Incluir placa
            pyautogui.click(76, 535, duration=0.2)
            pyautogui.click(469, 276, duration=0.2)
            # Remove o caractere "¨" da string PLACA
            PLACA = PLACA.replace("¨", "")
            pyautogui.click(534,275,duration=0.2)
            pyautogui.click(561,177,duration=0.2)
            pyautogui.click(464,245,duration=0.2)
            pyautogui.click(656,179,duration=0.2)
            PLACA = PLACA.replace("¨", "")
            pyautogui.typewrite(PLACA)
            PLACA = PLACA.replace("¨", "")
            pyautogui.click(956,179,duration=0.2)
            pyautogui.click(834,611,duration=0.2)
            #pyautogui.write(data)
            keyboard.press('tab')
            pyautogui.click(762, 541, duration=0.2)
            pyautogui.click(601, 535, duration=0.2)

            # Aba itens
            pyautogui.click(129, 98, duration=0.2)
            pyautogui.click(65, 287, duration=0.2)
            pyautogui.click(205, 131, duration=0.2)
            pyautogui.click(538, 177, duration=0.2)
            pyautogui.click(474, 231, duration=0.2)

            # Filtro
            pyautogui.click(615, 177, duration=0.2)
            pyautogui.write(PRODUTO)
            pyautogui.click(607, 180, duration=0.2)
            keyboard.press('backspace')
            pyautogui.click(959, 177, duration=0.2)
            pyautogui.click(719, 240, duration=0.2)
            pyautogui.click(832, 614, duration=0.2)

            # Quantidade e total
            pyautogui.click(170, 193, duration=0.2)
            pyautogui.typewrite('1')
            pyautogui.click(567, 194, duration=0.2)
            pyautogui.write(str(TOTAL))
            keyboard.press('tab')

            # Tipo de compra
            pyautogui.click(520, 250, duration=0.2)
            pyautogui.write('s')
            keyboard.press('tab')
            
            # Encerrar processo
            pyautogui.click(318, 281, duration=0.2)
            pyautogui.click(1096,702,duration=0.2)
            sleep(2)

            # Tempo inicial para a verificação da imagem
            tempo_inicial_verificacao = time()

            # Verificar se a imagem está na tela
            try:
                while time() - tempo_inicial_verificacao < tempo_max_espera:
                    if pyautogui.locateOnScreen("confirmar.png", confidence=0.8) is not None:                        
                        # Se a imagem for encontrada, execute o último click e saia do loop
                        pyautogui.click(883,421, duration=0.2)  
                        sleep(1)

                        # Adiciona a linha processada ao arquivo de registro
                        linhas_processadas.append(row_number)
                        with open('linhas_processadas.txt', 'w') as f:
                            for linha in linhas_processadas:
                                f.write(str(linha) + '\n')

                        # Zera o contador de tentativas para esta linha
                        tentativas_por_linha.pop(row_number, None)

                        # Avança para a próxima linha
                        row_number += 1
                        break
                    sleep(1)
                else:
                    print(f"Imagem não encontrada na linha {row_number}.")
                    # Verificar o número de tentativas para esta linha
                    if row_number in tentativas_por_linha:
                        tentativas_por_linha[row_number] += 1
                    else:
                        tentativas_por_linha[row_number] = 1

                    # Se a linha foi tentada 3 vezes, armazenar os dados em um arquivo de texto
                    if tentativas_por_linha[row_number] == 1:
                        with open('linhas_nao_processadas.txt', 'a') as f:
                            f.write(f"Dados da linha {row_number}: {TITULO}, {CUSTO}, {AREA}, {TABELA}, {PLACA}, {PRODUTO}, {TOTAL}, {DATA}\n")

                    # Encerra o processo atual
                    subprocess.run(['taskkill', '/f', '/im', 'javaw.exe'])
                    sleep(10)  # Espera 5 segundos
                    # Iniciar o KMM
                    subprocess.run(['"C:\\Users\\aprendiz.pcm\\Desktop\\site\\abrirKMM.bat"'])
            except pyautogui.ImageNotFoundException:
                print(f"Imagem não encontrada na linha {row_number}.")
                # Verificar o número de tentativas para esta linha
                if row_number in tentativas_por_linha:
                    tentativas_por_linha[row_number] += 1
                else:
                    tentativas_por_linha[row_number] = 1

                # Se a linha foi tentada 3 vezes, armazenar os dados em um arquivo de texto
            if tentativas_por_linha.get(row_number, 0) == 1:
                with open('linhas_nao_processadas.txt', 'a') as f:
                    f.write(f"Dados da linha {row_number}: {TITULO}, {CUSTO}, {AREA}, {TABELA}, {PLACA}, {PRODUTO}, {TOTAL}, {DATA}\n")

                # Encerra o processo atual
                subprocess.run(['taskkill', '/f', '/im', 'javaw.exe'])
                sleep(10)  # Espera 5 segundos
                # Iniciar o KMM
                keyboard.press_and_release('win+r')
                sleep(5)
                keyboard.press('delete')
                sleep(5)
                pyautogui.write('cmd')
                sleep(5)
                keyboard.press('enter')
                sleep(5)
                pyautogui.write('start C:\\Users\\Admin.Mix\\Desktop\\KMM.lnk')
                sleep(5)
                keyboard.press('enter')
                sleep(5)
                subprocess.run(['taskkill', '/f', '/im', 'cmd.exe'])
                pyautogui.click(708,399,duration=5)
                pyautogui.doubleClick(708,399,duration=5)
                keyboard.press('delete')
                pyautogui.write('ANTONIO.JUNIOR')
                pyautogui.click(602,427,duration=5)
                pyautogui.write('18117702')
                sleep(5)
                keyboard.press('enter')
                sleep(5)
                window_title = "KMM Corporativo :: 76.667.682 - BUDEL TRANSPORTES LTDA v3.23.1 Corporate"
                bring_window_to_front_and_maximize(window_title)
                sleep(5)
                pyautogui.write('csc')
                sleep(5)
                keyboard.press('enter')
                sleep(5)
                pyautogui.click(1113,15,duration=5)
                row_number += 1
        else:
            # Avançar para a próxima linha se a linha atual não tiver elementos suficientes
                row_number += 1

    # Salvar as alterações no arquivo Excel
    workbook.save(caminho_arquivo_excel)
    # Fechar o arquivo Excel após terminar de usar
    workbook.close()

# Chamar a função para iniciar o código
iniciar_codigo()
